
import os
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import Progressbar
from shutil import copyfile
import openpyxl
import time
import datetime
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
from PyPDF2 import PdfFileReader
from fuzzywuzzy import process, fuzz
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

window = Tk()


def browse_button():
    global pddm_source_file
    pddm_source_file = filedialog.askdirectory()
    source_name.set(pddm_source_file)
    print('FOLDER SELECTED IS ', pddm_source_file)


def close_window():
    answer = messagebox.askyesno("Cancel  APP", "Are you sure to cancel the app?")
    if answer is True:
        window.destroy()


style = ttk.Style()
style.theme_use('default')
style.configure("black.Horizontal.TProgressbar", background='green')

style = ttk.Style()
style.theme_use('default')
style.configure("black.Horizontal.TProgressbar", background='green')

bar = Progressbar(window, length=790, style='black.Horizontal.TProgressbar')
bar.place(x=60, y=240)

window.title('Welcome to PDF Properties Reader || Developed By Purna || V1.0')
window.geometry("1324x450")


def extract_information():
    try:
        print(pddm_source_file)
        print('********************PROGRAM STARTED TO READ PDF FILE*****************************')
        file_location = r'' + pddm_source_file
        file_path_val = os.listdir(file_location)
        print(file_location)

        dir_only_name = pddm_source_file
        current_date = datetime.datetime.today().strftime('%d-%b-%Y')
        current_date_folder = str(current_date)

        folder_with_date = 'OUTPUT FILE' + ' ' + current_date_folder
        generated_folder = os.path.join(dir_only_name, folder_with_date)

        # # create 'dynamic' dir, if it does not exist
        if not os.path.exists(generated_folder):
            os.makedirs(generated_folder)

        file_name = generated_folder + '/' + 'Device Output.xlsx'
        workbook = Workbook()
        sheet = workbook.active

        sheet["A1"] = "File Name"
        sheet["B1"] = "Application Name"
        log_sheet = workbook.create_sheet("Log_sheet")
        log_sheet.title = "Log_sheet"
        file_name_log = log_sheet.cell(row=1,column=1)
        file_name_log.value = 'File name'
        arr_len = len(file_path_val)

        cnt_prgs = 0
        cnt_log =0
        for j, crs_file_name in enumerate(file_path_val):
            extension = os.path.splitext(crs_file_name)[1]
            # print(extension)
            # if crs_file_name.endswith(('.pdf','.PDF')):
            if extension.lower() == '.pdf':
                comment_file = file_location + '/' + crs_file_name
                # print(comment_file)
                total_cnt = arr_len

                try:
                    with open(comment_file, 'rb') as f:
                        pdf = PdfFileReader(f)
                        information = pdf.getDocumentInfo()

                        file_value = sheet.cell(row=cnt_prgs + 2, column=1)
                        file_value.value = crs_file_name

                        application_name = sheet.cell(row=cnt_prgs + 2, column=2)
                        if information is not None:
                            if information.creator is None:
                                application_name.value = 'N/A'
                            elif information.creator == '':
                                application_name.value = 'N/A'
                            else:
                                application_name.value = information.creator
                                # print(information.creator
                        cnt_prgs += 1
                        print(cnt_prgs, 'PDF READ SUCCESSFULLY OUT OF', total_cnt)

                except:
                    cnt_log += 1
                    file_name_log = log_sheet.cell(row=cnt_log+1,column=1)
                    file_name_log.value = crs_file_name
                    print(cnt_log,'This is error file',crs_file_name)
                    pass

        workbook.save(filename=file_name)

        print("\n******************* PROGRAM EXECUTED SUCCESSFULLY.**********************")

        answer = messagebox.askyesno("Success Information",
                                     "Program run successfully. Thank you for using the program. Press Yes to Close")
        if answer is True:
            window.destroy()
    except Exception as ex:
        print("Could not open file! Please close !")
        pass
        messagebox.showerror("Error", ex)


source_lbl = Label(window, text="SOURCE FOLDER", fg='black', bg="grey", font=("Helvetica", 13))
source_lbl.place(x=60, y=50)

source_name = StringVar(None)
source_txtfld = Entry(window, textvariable=source_name, bd=2, width=74)
source_txtfld.place(x=220, y=50)

browse = Button(window, text="BROWSE SOURCE PDF FOLDER", fg='BLACK', bg="darkgrey", command=browse_button)
browse.place(x=700, y=40)
browse.config(width=40, height=2)

submit_btn = Button(window, text="SUBMIT", bg='#006400', fg='white', command=extract_information)
submit_btn.place(x=60, y=100)
submit_btn.config(width=20, height=3)

cancel = Button(window, text="CANCEL", command=close_window, bg='#8B0000', fg='white')
cancel.place(x=230, y=100)
cancel.config(width=20, height=3)

note_lbl = Label(window, text="Note:Please close all file before Executing the program.", fg='darkred', bg="grey",
                 font=("Helvetica", 13))
note_lbl.place(x=60, y=180)

note_lbl_name = Label(window, text="Script written By Purna ", fg='darkred', bg="grey",
                      font=("Helvetica", 14))
note_lbl_name.place(x=60, y=380)

# FILE COPIER END
windowWidth = window.winfo_reqwidth()
windowHeight = window.winfo_reqheight()
window.configure(bg='grey')
# Gets both half the screen width/height and window width/height
positionRight = int(window.winfo_screenwidth() / 2 - windowWidth / 2)
positionDown = int(window.winfo_screenheight() / 2 - windowHeight / 2)
window.geometry("+{}+{}".format(positionRight, positionDown))
window.resizable(0, 0)
window.mainloop()
